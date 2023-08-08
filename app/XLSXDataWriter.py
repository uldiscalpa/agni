import openpyxl


class XLSXDataWriter:
    def __init__(self, template_path, data_row_index: int, header_row_index: int = 1, ):
        self.template_path = template_path
        self.workbook = None
        self.sheet = None
        self.data_row_index = data_row_index
        self.header_row_index = header_row_index
        self.data_row_styles = None

    def load_template(self):
        self.workbook = openpyxl.load_workbook(self.template_path)
        self.sheet = self.workbook.active
        self.find_data_row_styles()

    def get_template_column_headers(self):
        # Assuming that the column headers are in the first row of the sheet
        self.column_headers = [
            cell.value for cell in self.sheet[self.header_row_index]]

    def get_column_mapping(self, header_row):
        # Create a mapping of column names to column indices
        column_mapping = {}
        for col_idx, cell in enumerate(header_row, start=1):
            column_name = cell.value
            if column_name:
                column_mapping[column_name] = col_idx
        return column_mapping

    def find_data_row_styles(self):
        # Find the data row index and save its cell styles
        if not self.data_row_index:
            for row in self.sheet.iter_rows(values_only=True):
                if all(cell is not None for cell in row):
                    self.data_row_styles = [cell._style for cell in row]
                    break
        self.data_row_styles = [
            cell._style for cell in self.sheet[self.data_row_index]]

    def find_first_empty_row(self):
        # Find the first empty row after the data row
        if self.data_row_styles is None:
            self.find_data_row_styles()
        max_row = self.sheet.max_row
        for row in range(max_row, 1, -1):
            if all(not self.sheet.cell(row=row, column=col).value for col in range(1, self.sheet.max_column + 1)):
                return row + 1
        return max_row + 1

    def apply_style(self, new_row_count):
        # Apply styles to the newly added rows
        if self.data_row_styles:
            new_row = self.sheet[self.sheet.max_row]
            for cell, style in zip(new_row, self.data_row_styles):
                cell._style = style

    def append_data(self, data):
        self.load_template()
        self.sheet.delete_rows(self.data_row_index)
        header_row = self.sheet[self.header_row_index]

        # Get the column mapping for header columns
        column_mapping = self.get_column_mapping(header_row)

        data_headers = data[0]
        # Create a new data row with None for all columns
        for row_data in data[1:]:
            new_data_row = [None] * len(header_row)
            for header, data_value in zip(data_headers, row_data):
                col_idx = column_mapping.get(header, None)
                if col_idx is not None:
                    new_data_row[col_idx - 1] = data_value
            self.sheet.append(new_data_row)
            if self.data_row_styles:
                new_row = self.sheet[self.sheet.max_row]
                for cell, _style in zip(new_row, self.data_row_styles):
                    cell._style = _style

    def save(self, output_file):
        self.workbook.save(output_file)

# Example usage of XLSXDataWriter class


def remove_duplicates(input_list):
    unique_lists = []
    seen = set()

    for sublist in input_list:
        tuple_sublist = tuple(sublist)
        if tuple_sublist not in seen:
            unique_lists.append(sublist)
            seen.add(tuple_sublist)

    return unique_lists


def remove_duplicate_rows(input_file, output_file):
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    # Initialize a set to keep track of unique rows
    unique_rows = set()

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Convert the row data to a tuple to make it hashable
        row_tuple = tuple(row)

        # Check if the row is not a duplicate
        if row_tuple not in unique_rows:
            unique_rows.add(row_tuple)

    # Create a new workbook and sheet to write the unique data
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Write the unique rows to the new sheet
    for row_data in unique_rows:
        new_sheet.append(row_data)

    # Save the new workbook to the output file
    new_workbook.save(output_file)


def main():
    template_file = "..template.xlsx"

    # Sample data to be inserted into the XLSX file
    data_to_insert = [
        ["New Data 1", 300, 400],
        ["New Data 2", 500, 600],
        # Add more rows as needed
    ]

    writer = XLSXDataWriter(template_file)
    writer.append_data(data_to_insert)


if __name__ == "__main__":
    main()
