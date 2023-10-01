import pandas as pd
from typing import List
from openpyxl import load_workbook


class ExcelHandler:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def fetch_local_data(self) -> pd.DataFrame:
        """
        Reads the local Excel file and returns its contents as a DataFrame.
        """
        return pd.read_excel(self.file_path)

    def adjust_columns_width(self, file_path: str = None) -> None:
        """
        Adjusts the width of the columns in all sheets of the Excel file to fit the contents.
        """
        # Read the Excel file into a DataFrame
        file = file_path if file_path else self.file_path
        df = pd.read_excel(file)

        # Get the maximum length of the values in each column for all sheets
        max_lengths = {}
        for sheet_name in pd.ExcelFile(file).sheet_names:
            sheet_df = pd.read_excel(file, sheet_name=sheet_name)
            max_lengths[sheet_name] = [
                len(str(sheet_df[column_name].astype(str).max())) for column_name in sheet_df.columns]

        # Adjust the width of each column in all sheets to fit the contents
        workbook = load_workbook(file)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for i, column_name in enumerate(df.columns):
                worksheet.column_dimensions[worksheet.cell(
                    row=1, column=i+1).column_letter].width = max_lengths[sheet_name][i] + 2
        workbook.save(file)

    def read_column_values(self, column_name: List[str], sheet_name: str = 'Sheet1') -> List[List]:
        """
        Reads the Excel file and extracts the values from the specified column.
        Returns a list of unique values from the specified column, sorted in descending order.
        """
        # Read the Excel file into a DataFrame
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)

        # Get the values from the specified column and sort them in descending order
        column_values = df[column_name].drop_duplicates().values.tolist()

        return column_values

    def write_values_to_excel(self, values: List[List], file_name: str, sheet_name: str = "Sheet 1") -> None:
        """
        Writes the specified values to a new sheet in an Excel file, adjusts the column widths, and saves the file with the specified name.
        :param values: A list of lists containing the values to write to the Excel file.
        :param sheet_name: The name of the sheet to write the values to.
        :param file_name: The name of the file to save the Excel file as.
        """
        # Create a DataFrame from the values and write it to a new sheet in the Excel file
        df = pd.DataFrame(values)
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        writer.book = load_workbook(self.file_path)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()

        # Adjust the column widths in the new sheet
        self.adjust_columns_width()

    def write_data_frame_to_excel(self, df: pd.DataFrame, file_name: str, sheet_name: str = "Sheet 1") -> None:
        """
        Writes the specified DataFrame to a new sheet in an Excel file, adjusts the column widths, and saves the file with the specified name.
        :param df: The DataFrame to write to the Excel file.
        :param sheet_name: The name of the sheet to write the values to.
        :param file_name: The name of the file to save the Excel file as.
        """
        # Write the DataFrame to a new sheet in the Excel file
        df.to_excel(file_name, sheet_name=sheet_name,
                    index=False, engine='openpyxl')
        # Adjust the column widths in the new sheet
        self.adjust_columns_width()

    def read_data(self) -> pd.DataFrame:
        """
        Reads the Excel file and returns its contents as a DataFrame.
        """
        return pd.read_excel(self.file_path)

    @staticmethod
    def write_values_to_excel_staticmethod(file_path: str, values: List[List[str]], sheet_name: str) -> None:
        """
        Writes the specified values to a new sheet in a new Excel file, adjusts the column widths, and saves the file with the specified name.
        This is a static method that does not depend on the state of the ExcelHandler class.
        :param file_path: The path to the Excel file to read the data from.
        :param values: A list of lists containing the values to write to the new Excel file.
        :param sheet_name: The name of the sheet to write the values to.
        :param file_name: The name of the new Excel file to save the data to.
        """
        # Create a DataFrame from the values and write it to a new sheet in a new Excel file
        # Extract the column names and data from values
        columns = values[0]
        data = values[1:]

        # Create a DataFrame from the data and column names
        df = pd.DataFrame(data, columns=columns)
        df = df.drop("index", axis=1, errors="ignore")
        df.reset_index(drop=True)
        df.to_excel(file_path, sheet_name=sheet_name,
                    index=False, engine='openpyxl')

    @staticmethod
    def adjust_columns_width_staticmethod(file_path: str = None) -> None:
        """
        Adjusts the width of the columns in all sheets of the Excel file to fit the contents.
        """
        # Read the Excel file into a DataFrame
        file = file_path if file_path else file_path
        df = pd.read_excel(file,  engine='openpyxl')

        # Get the maximum length of the values in each column for all sheets
        max_lengths = {}
        for sheet_name in pd.ExcelFile(file).sheet_names:
            sheet_df = pd.read_excel(file, sheet_name=sheet_name)
            max_lengths[sheet_name] = [
                len(str(sheet_df[column_name].astype(str).max())) for column_name in sheet_df.columns]

        # Adjust the width of each column in all sheets to fit the contents
        workbook = load_workbook(file)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for i, column_name in enumerate(df.columns):
                worksheet.column_dimensions[worksheet.cell(
                    row=1, column=i+1).column_letter].width = max_lengths[sheet_name][i] + 2
        workbook.save(file)


def main() -> None:
    file_path = "..\\data\\projects.xlsx"
    column_name = "code"
    excel_reader = ExcelHandler(file_path, column_name)
    values = excel_reader.read_column_values()


if __name__ == "__main__":
    main()
